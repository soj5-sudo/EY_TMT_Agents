from __future__ import annotations

import io
import json
import re
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent

TRACKER_ID = "17EKJTmW0_eTIbS9ShxWi9Ym4A4_pIZRB9FBV2Tq5Dx4"
PEERS_ID = "1GCmWWKcNa4VTt86gJu43QRv7T2WT3DhcNaCb2mJYeXI"

UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/126.0 Safari/537.36"}

def fetch(sheet_id: str) -> openpyxl.Workbook:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=120) as r:
        data = r.read()
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)

def text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def num(v):
    if v is None or isinstance(v, str):
        return None
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None

def clean_name(v: str) -> str:
        return re.sub(r"\s*\d?$", "", text(v)).replace(" Limited", "").replace(" Technology", "").strip()

METRIC_SHEETS = [
    ("Revenue", "revenueUsdM", 1, 2, "USD m"),
    ("Hiring", "headcountK", 1, 2, "thousands"),
    ("Attrition", "attritionPct", 1, 2, "share"),
    ("Cap Utilization", "utilisationPct", 1, 2, "share"),
]

def find_header(rows) -> tuple[int, int] | tuple[None, None]:
        for i, r in enumerate(rows[:40]):
        for j, c in enumerate(r[:6]):
            if text(c).lower().startswith("companie"):
                return i, j
    return None, None

def cell(r, i: int) -> object:
    return r[i] if i < len(r) else None

def read_metric_sheet(wb, sheet: str, prior_col: int, latest_col: int):
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header, base = find_header(rows)
    if header is None:
        return {}, ("", "")

    periods = (text(cell(rows[header], base + prior_col)), text(cell(rows[header], base + latest_col)))
    out: dict[str, dict] = {}

    for r in rows[header + 1 :]:
        name = clean_name(cell(r, base))
        if not name or name.lower().startswith(("note", "source", "1.")):
            break
        out[name] = {
            "prior": num(cell(r, base + prior_col)),
            "latest": num(cell(r, base + latest_col)),
            "status": text(cell(r, base + 3)),
            "reason": text(cell(r, base + 4))[:900],
            "source": text(cell(r, base + 5))[:200],
        }
    return out, periods

def read_margins(wb):
        ws = wb["EBIT and EBITDA margins"]
    rows = list(ws.iter_rows(values_only=True))
    header, base = find_header(rows)
    if header is None:
        return {}, ("", "")

    periods = (text(cell(rows[header], base + 1)), text(cell(rows[header], base + 2)))
    out: dict[str, dict] = {}

    for r in rows[header + 1 :]:
        name = clean_name(cell(r, base))
        if not name or name.lower().startswith(("note", "source", "1.")):
            break
        out[name] = {
            "ebit": {"prior": num(cell(r, base + 1)), "latest": num(cell(r, base + 2))},
            "ebitda": {"prior": num(cell(r, base + 3)), "latest": num(cell(r, base + 4))},
            "status": text(cell(r, base + 5)),
            "reason": text(cell(r, base + 6))[:900],
            "source": text(cell(r, base + 7))[:200],
        }
    return out, periods

def read_narrative(wb, sheet: str, limit: int = 60):
        if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    out = []
    for r in ws.iter_rows(values_only=True):
        cells = [text(c) for c in r[:8]]
        if not any(cells):
            continue
        if len(cells[0]) > 1 and any(len(c) > 25 for c in cells[1:]):
            out.append([c[:700] for c in cells if c])
        if len(out) >= limit:
            break
    return out

VERTICALS = {
    "IT Services": "IT services",
    "E R&D": "Engineering R and D",
    "Software": "Enterprise software",
    "Data Analytics": "Data analytics",
    "Contract centre-BPO KPO": "Contact centre",
}

def read_peers(wb):
    out = []
    for sheet, vertical in VERTICALS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = next(
            (i for i, r in enumerate(rows) if text(r[0]).lower().startswith("currency")), None
        )
        if header is None:
            continue

        cols = [text(c).lower() for c in rows[header]]

        def col(*names, default=None):
            for n in names:
                for i, c in enumerate(cols):
                    if c.startswith(n):
                        return i
            return default

        i_hq = col("headquarter")
        i_service = col("service")
        i_mcap = col("mcap")
        i_fs = col("last reported")
        i_gaap = col("gaap")
        i_rev = col("revenue")
        i_ebitda = col("ebitda (")
        i_pct = col("ebitda %")
        i_inv = col("key investor")
        i_tyear = col("last transaction year")
        i_tdesc = col("last transaction", "transaction detail")

        for r in rows[header + 1 :]:
            name = text(r[0])
            if not name or name.lower().startswith(("note", "source")):
                continue
            rev = num(r[i_rev]) if i_rev is not None else None
            if rev is None and num(r[i_mcap] if i_mcap is not None else None) is None:
                continue
            out.append(
                {
                    "name": name,
                    "vertical": vertical,
                    "headquarters": text(r[i_hq]) if i_hq is not None else "",
                    "service": text(r[i_service]) if i_service is not None else "",
                    "marketCapUsdM": num(r[i_mcap]) if i_mcap is not None else None,
                    "lastReported": text(r[i_fs]) if i_fs is not None else "",
                    "gaap": text(r[i_gaap]) if i_gaap is not None else "",
                    "revenueUsdM": rev,
                    "ebitdaUsdM": num(r[i_ebitda]) if i_ebitda is not None else None,
                    "ebitdaPct": num(r[i_pct]) if i_pct is not None else None,
                    "keyInvestors": (text(r[i_inv]) if i_inv is not None else "")[:260],
                    "lastTransactionYear": (text(r[i_tyear]) if i_tyear is not None else "")[:12],
                    "lastTransaction": (text(r[i_tdesc]) if i_tdesc is not None else "")[:400],
                }
            )
    return out

TIER1 = {"TCS", "Infosys", "Wipro", "HCL Tech", "Tech Mahindra"}

def main() -> int:
    print("fetching the quarterly tracker")
    tracker = fetch(TRACKER_ID)
    print("fetching the peer universe")
    peers_wb = fetch(PEERS_ID)

    companies: dict[str, dict] = {}
    periods = ("", "")

    for sheet, key, pc, lc, unit in METRIC_SHEETS:
        if sheet not in tracker.sheetnames:
            continue
        data, per = read_metric_sheet(tracker, sheet, pc, lc)
        if per[1]:
            periods = per
        for name, v in data.items():
            companies.setdefault(name, {"name": name, "metrics": {}})
            companies[name]["metrics"][key] = {**v, "unit": unit}
        print(f"  {sheet:22} {len(data):2} companies  periods {per[0]} to {per[1]}")

    margins, mper = read_margins(tracker)
    for name, v in margins.items():
        companies.setdefault(name, {"name": name, "metrics": {}})
        companies[name]["metrics"]["ebitMarginPct"] = {
            "prior": v["ebit"]["prior"],
            "latest": v["ebit"]["latest"],
            "status": v["status"],
            "reason": v["reason"],
            "source": v["source"],
            "unit": "share",
        }
        if v["ebitda"]["latest"] is not None:
            companies[name]["metrics"]["ebitdaMarginPct"] = {
                "prior": v["ebitda"]["prior"],
                "latest": v["ebitda"]["latest"],
                "status": v["status"],
                "reason": "",
                "source": v["source"],
                "unit": "share",
            }
    print(f"  {'EBIT and EBITDA':22} {len(margins):2} companies  periods {mper[0]} to {mper[1]}")

    rows = []
    for name, c in companies.items():
        rows.append(
            {
                "name": name,
                "tier": "Indian tier-1" if name in TIER1 else "Indian mid-tier",
                "metrics": c["metrics"],
            }
        )
    rows.sort(key=lambda r: (r["tier"] != "Indian tier-1", r["name"]))

    narratives = {
        "aiAndGenAi": read_narrative(tracker, "Inv in AI and Gen AI"),
        "mergersAndAcquisitions": read_narrative(tracker, "M&A deals"),
        "geographicExpansion": read_narrative(tracker, "Geographic expansion"),
        "managementAnnouncements": read_narrative(tracker, "Management announcements"),
        "revenueGuidance": read_narrative(tracker, "Revenue Guidance"),
        "marginGuidance": read_narrative(tracker, "Margin Guidance"),
    }
    for k, v in narratives.items():
        print(f"  {k:22} {len(v)} entries")

    peers = read_peers(peers_wb)
    by_vertical: dict[str, int] = {}
    for p in peers:
        by_vertical[p["vertical"]] = by_vertical.get(p["vertical"], 0) + 1
    print(f"  peer universe          {len(peers)} companies  {by_vertical}")

    taken = datetime.now(timezone.utc).isoformat()

    out = REPO / "lib" / "data" / "sector-tracker.ts"
    out.write_text(
        '/**\n'
        ' * Quarterly sector tracker.\n'
        ' *\n'
        ' * Generated by scraper/ingest_sheets from two published workbooks, both\n'
        ' * shared for public read and exported over a plain request with no key.\n'
        ' *\n'
        ' * This is not the same material as the rest of the console, and is treated\n'
        ' * differently. The dashboards compute their figures from filings on every\n'
        ' * request; this is a curated quarterly record carrying what management said\n'
        ' * about a quarter and why a measure moved, which no filing publishes in a\n'
        ' * machine readable form. Every cell keeps its own source line, and the\n'
        ' * dataset is stamped with the date it was read.\n'
        ' *\n'
        ' * Refresh with: npm run ingest:sheets\n'
        ' */\n\n'
        "export interface TrackerMeasure {\n"
        "  prior: number | null;\n"
        "  latest: number | null;\n"
        "  /** Movement as the tracker states it, for example \"-1.7% QoQ; 3.5% YoY\". */\n"
        "  status: string;\n"
        "  /** What management gave as the reason. Verbatim from the tracker. */\n"
        "  reason: string;\n"
        "  /** Where the reason came from: a transcript, an analyst report, an article. */\n"
        "  source: string;\n"
        "  unit: string;\n"
        "}\n\n"
        "export interface TrackerCompany {\n"
        "  name: string;\n"
        '  tier: "Indian tier-1" | "Indian mid-tier";\n'
        "  metrics: Record<string, TrackerMeasure>;\n"
        "}\n\n"
        f"/** The two quarters the tracker compares. */\nexport const TRACKER_PERIODS = {json.dumps({'prior': periods[0], 'latest': periods[1]})};\n\n"
        f"export const TRACKER_TAKEN_AT = {json.dumps(taken)};\n\n"
        f"export const TRACKER: TrackerCompany[] = {json.dumps(rows, indent=1)};\n\n"
        "/** Prose sections of the tracker, kept as rows of cells. */\n"
        f"export const TRACKER_NARRATIVE: Record<string, string[][]> = {json.dumps(narratives, indent=1)};\n"
    )

    out2 = REPO / "lib" / "data" / "peer-universe.ts"
    out2.write_text(
        '/**\n'
        ' * The wider technology services peer universe.\n'
        ' *\n'
        ' * Generated by scraper/ingest_sheets from a published workbook. Listed\n'
        ' * companies across five verticals, with the reported figures, the GAAP each\n'
        ' * one reports under, its principal holders and its last transaction.\n'
        ' *\n'
        ' * The figures here are as of the period each row names, which differs by\n'
        ' * company and is recorded per row rather than assumed. Where the console can\n'
        ' * compute a current figure from a filing it does that instead; this is the\n'
        ' * comparison set, not the live record.\n'
        ' *\n'
        ' * Refresh with: npm run ingest:sheets\n'
        ' */\n\n'
        "export interface PeerCompany {\n"
        "  name: string;\n"
        "  vertical: string;\n"
        "  headquarters: string;\n"
        "  service: string;\n"
        "  marketCapUsdM: number | null;\n"
        "  /** The period the figures on this row are taken from. */\n"
        "  lastReported: string;\n"
        "  gaap: string;\n"
        "  revenueUsdM: number | null;\n"
        "  ebitdaUsdM: number | null;\n"
        "  ebitdaPct: number | null;\n"
        "  keyInvestors: string;\n"
        "  lastTransactionYear: string;\n"
        "  lastTransaction: string;\n"
        "}\n\n"
        f"export const PEER_TAKEN_AT = {json.dumps(taken)};\n\n"
        f"export const PEER_UNIVERSE: PeerCompany[] = {json.dumps(peers, indent=1)};\n"
    )

    print(f"\nwrote lib/data/sector-tracker.ts: {len(rows)} companies")
    print(f"wrote lib/data/peer-universe.ts: {len(peers)} companies")
    return 0

if __name__ == "__main__":
    sys.exit(main())
